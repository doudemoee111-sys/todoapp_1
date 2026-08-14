#!/usr/bin/env python3
"""蓄積データ（基準CSV + 日次差分JSON）を、元ワークブックと同じ列構成の .xlsx に書き出す。

  python3 fx/scripts/export_excel.py [出力パス]

6銘柄の各シートを、元の列構成で生成する:
  日付 / 始値 / 高値 / 安値 / 終値 / 高値ー安値 / 始値ー終値 / 始値ー安値 / 始値ー高値 / 年 / 月 / 日 / 曜日 / 第N週
（統計分析「集計用」シートは Web ダッシュボードが担うため、ここではデータ6シートのみ）
"""
import csv, json, os, sys, datetime
from pathlib import Path

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("openpyxl が必要です: pip install openpyxl")

ROOT = Path(__file__).resolve().parents[1]
DATA = ROOT / "data"
DAILY = DATA / "daily"

# key -> (シート名, 小数桁)
SYMBOLS = [
    ("USDJPY", "ドル円", 3), ("EURJPY", "ユーロ円", 3), ("GBPJPY", "ポンド円", 3),
    ("EURUSD", "ユーロドル", 5), ("GBPUSD", "ポンドドル", 5), ("GOLD", "ゴールド", 2),
]
WD = ["月", "火", "水", "木", "金", "土", "日"]  # Python weekday(): Mon=0
HEADER = ["日付", "始値", "高値", "安値", "終値", "高値ー安値", "始値ー終値",
          "始値ー安値", "始値ー高値", "年", "月", "日", "曜日", "第N週"]


def load_base(key):
    p = DATA / f"{key}.csv"
    rows = {}
    if p.exists():
        with open(p, newline="") as f:
            r = csv.reader(f)
            next(r, None)
            for row in r:
                if len(row) < 5:
                    continue
                d = row[0]
                try:
                    rows[d] = (float(row[1]), float(row[2]), float(row[3]), float(row[4]))
                except ValueError:
                    continue
    return rows


def apply_deltas(rows, key):
    if not DAILY.exists():
        return
    for fp in sorted(DAILY.glob("*.json")):
        try:
            obj = json.loads(fp.read_text())
        except Exception:
            continue
        date = obj.get("date") or fp.stem
        bar = obj.get(key)
        if not bar:
            continue
        try:
            rows[date] = (float(bar["open"]), float(bar["high"]), float(bar["low"]), float(bar["close"]))
        except (KeyError, ValueError, TypeError):
            continue


def main():
    out = Path(sys.argv[1]) if len(sys.argv) > 1 else (ROOT / "FX_蓄積データ.xlsx")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    total = 0
    for key, sheet, dec in SYMBOLS:
        rows = load_base(key)
        apply_deltas(rows, key)
        ws = wb.create_sheet(title=sheet)
        ws.append(HEADER)
        for d in sorted(rows):
            o, h, l, c = rows[d]
            dt = datetime.date.fromisoformat(d)
            rnd = lambda v: round(v, dec)
            ws.append([
                dt, rnd(o), rnd(h), rnd(l), rnd(c),
                rnd(h - l), rnd(o - c), rnd(o - l), rnd(o - h),
                dt.year, dt.month, dt.day, WD[dt.weekday()],
                (dt.day - 1) // 7 + 1,
            ])
        # 日付列の表示形式 + オートフィルタ + ヘッダ固定
        for cell in ws["A"][1:]:
            cell.number_format = "yyyy-mm-dd"
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADER))}{ws.max_row}"
        ws.column_dimensions["A"].width = 12
        n = ws.max_row - 1
        total += n
        print(f"{key:7} -> シート「{sheet}」 {n} 行 ({sorted(rows)[0]}..{sorted(rows)[-1]})")
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(f"保存: {out}  （計 {total} 行 / {len(SYMBOLS)} シート）")


if __name__ == "__main__":
    main()
